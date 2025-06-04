#!/usr/bin/env python
# coding=utf-8
import os
import pandas as pd
import threading
import time
import random
import h5py
import numpy as np
import torch
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import copy

# --- Configuration ---
_chunk_size = 500  # Save every 10,000 samples
_delete_chunks_after_merge = True  # Set to False to keep intermediate chunk files

# --- Global State (Protected by Lock) ---
_buffer = []
_chunk_files = []
_samples_in_buffer = 0
_total_samples_logged_this_epoch = 0
_current_epoch = None
_file_lock = threading.Lock()

# Define headers based on the original structure
_headers = [
    "Index", "Sample Index", "stats/Loss_total", "stats_IoU", "Seq Name",
    "Template Frame ID", "Template Frame Path", "Search Frame ID", "Seq ID",
    "Seq Path", "Class Name", "Vid ID", "Search Names", "Search Path"
]


# --- Filename Generation ---
def _get_chunk_filename(epoch, start_index, end_index):
    # Save in the root directory where the script is run
    return f'samples_log_epoch_{epoch}_sample_{start_index}_{end_index}.xlsx'


def _get_final_filename(epoch, total_samples):
    # Save in the root directory where the script is run
    return f'samples_log_epoch_{epoch}_all_sample_1_{total_samples}.xlsx'


# --- Helper Functions ---
def _safe_str_list(value):
    """Safely convert lists or other types to string."""
    if isinstance(value, list):
        # Handle nested lists if necessary, assuming simple list of strings/numbers for now
        return ", ".join(map(str, value))
    elif value is None:
        return ""
    else:
        return str(value)


def _format_excel_file(filename):
    """Applies basic formatting (alignment, column width) to an Excel file."""
    try:
        from openpyxl import load_workbook  # Import locally to avoid dependency if not used
        wb = load_workbook(filename)
        ws = wb.active

        # Center alignment for all cells
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align

        # Auto-adjust column widths based on content (max length)
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            # Check header length first
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                max_length = len(str(header_cell.value))
            # Check content length (sample a few rows for efficiency if needed)
            for cell in column_cells[1:]:  # Skip header
                try:
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except:  # Handle potential errors with cell values
                    pass
            # Add padding
            adjusted_width = max_length + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Set header row height
        ws.row_dimensions[1].height = 25
        # Optionally set default row height for others if needed
        # for i in range(2, ws.max_row + 1):
        #     ws.row_dimensions[i].height = 20

        wb.save(filename)
        # print(f"Applied formatting to {filename}")
    except ImportError:
        print("Warning: openpyxl not found. Cannot apply Excel formatting.")
    except Exception as e:
        print(f"Error formatting Excel file {filename}: {e}")


# --- Core Logic ---
def _save_chunk(epoch, start_index, end_index, data_to_save):
    """Saves the current buffer to a chunk file."""
    global _chunk_files
    if not data_to_save:
        print("No data in buffer to save as chunk.")
        return

    filename = _get_chunk_filename(epoch, start_index, end_index)
    print(f"Saving chunk {start_index}-{end_index} for epoch {epoch} to {filename}...")

    try:
        df = pd.DataFrame(data_to_save)
        # Ensure columns are in the correct order
        df = df.reindex(columns=_headers)

        # Save using pandas, specifying the engine
        df.to_excel(filename, index=False, engine='openpyxl')

        # Apply formatting after saving
        _format_excel_file(filename)

        _chunk_files.append(filename)
        print(f"Successfully saved and formatted chunk: {filename}")

    except Exception as e:
        print(f"Error saving chunk {filename}: {e}")


def set_epoch(epoch_number):
    """Sets the current epoch, clearing buffers and state for the new epoch."""
    global _current_epoch, _buffer, _samples_in_buffer, _chunk_files, _total_samples_logged_this_epoch
    with _file_lock:
        if _current_epoch is not None and _current_epoch != epoch_number:
            # If finalize_epoch wasn't called by the trainer, call it defensively.
            print(
                f"Warning: Starting epoch {epoch_number} but previous epoch {_current_epoch} was not explicitly finalized. Finalizing {_current_epoch} now.")
            finalize_epoch(_current_epoch)

        print(f"Setting data recorder for epoch {epoch_number}. Clearing state.")
        _current_epoch = epoch_number
        _buffer = []
        _samples_in_buffer = 0
        _chunk_files = []
        _total_samples_logged_this_epoch = 0


def samples_stats_save(sample_index: int, data_info: dict, stats: dict):
    """
    Save sample statistics to the buffer for later logging to Excel.
    
    Args:
        sample_index: Index of the current sample
        data_info: Dictionary containing sample information
        stats: Dictionary containing sample statistics
    """
    global _buffer, _samples_in_buffer, _total_samples_logged_this_epoch

    # Determine epoch (should be set by trainer via set_epoch or passed in data_info)
    epoch = data_info.get('epoch', _current_epoch)
    if epoch is None:
        print("Error: Epoch not set in data_recorder. Cannot log data. Call set_epoch() first.")
        return
    # Ensure consistency if epoch changes unexpectedly mid-stream
    if epoch != _current_epoch:
        print(
            f"Warning: Logging data for epoch {epoch}, but recorder's current epoch is {_current_epoch}. Attempting to switch epoch.")
        set_epoch(epoch)

    # Prepare the data entry as a dictionary
    loss_total = stats.get("Loss/total", None)
    iou = stats.get("IoU", None)

    with _file_lock:
        _total_samples_logged_this_epoch += 1
        current_log_index = _total_samples_logged_this_epoch

        # Create the dictionary matching the headers
        log_entry = {
            "Index": current_log_index,
            "Sample Index": sample_index,
            "stats/Loss_total": loss_total,
            "stats_IoU": iou,
            "Seq Name": data_info.get("seq_name", ""),
            "Template Frame ID": _safe_str_list(data_info.get("template_ids")),
            "Template Frame Path": _safe_str_list(data_info.get("template_path")),
            "Search Frame ID": _safe_str_list(data_info.get("search_id")),
            "Seq ID": data_info.get("seq_id", ""),
            "Seq Path": data_info.get("seq_path", ""),
            "Class Name": data_info.get("class_name", ""),
            "Vid ID": data_info.get("vid_id", ""),
            "Search Names": _safe_str_list(data_info.get("search_names")),
            "Search Path": _safe_str_list(data_info.get("search_path"))
        }

        _buffer.append(log_entry)
        _samples_in_buffer += 1

        # Check if the buffer is full and needs to be saved as a chunk
        if _samples_in_buffer >= _chunk_size:
            start_index = current_log_index - _samples_in_buffer + 1
            end_index = current_log_index
            # Save the current buffer and clear it
            _save_chunk(epoch, start_index, end_index, _buffer)
            _buffer = []
            _samples_in_buffer = 0


def finalize_epoch(epoch):
    """Finalizes logging for the epoch: saves remaining buffer, merges chunks, cleans up."""
    global _buffer, _samples_in_buffer, _chunk_files, _total_samples_logged_this_epoch, _current_epoch

    with _file_lock:
        # Validate the epoch number
        if epoch is None:
            print("Error: Cannot finalize epoch, epoch number is None.")
            return
        if epoch != _current_epoch:
            print(
                f"Warning: Finalizing epoch {epoch}, but recorder's current epoch is {_current_epoch}. Finalization might use data from the wrong epoch if not careful.")
            # We proceed, assuming the caller knows which epoch to finalize.

        print(
            f"Finalizing data logging for epoch {epoch} (Total samples logged: {_total_samples_logged_this_epoch})...")

        # 1. Save any remaining data in the buffer as the last chunk
        if _samples_in_buffer > 0:
            start_index = _total_samples_logged_this_epoch - _samples_in_buffer + 1
            end_index = _total_samples_logged_this_epoch
            print(f"Saving final buffer chunk ({_samples_in_buffer} samples) for epoch {epoch}...")
            _save_chunk(epoch, start_index, end_index, _buffer)
            _buffer = []  # Clear buffer after saving
            _samples_in_buffer = 0
        else:
            print("No samples remaining in buffer.")

        # 2. Merge chunk files if any exist
        if not _chunk_files:
            print(f"No chunk files were created for epoch {epoch}. Nothing to merge.")
            # Reset state for the possibility of starting a new epoch later
            _current_epoch = None  # Mark as no longer active
            return

        print(f"Merging {_chunk_files} chunk files for epoch {epoch}...")
        all_data_frames = []
        for chunk_file in _chunk_files:
            try:
                print(f"Reading chunk file: {chunk_file}")
                df = pd.read_excel(chunk_file, engine='openpyxl')
                all_data_frames.append(df)
            except Exception as e:
                print(f"Error reading chunk file {chunk_file}: {e}. Skipping this chunk.")

        if not all_data_frames:
            print(
                f"Error: Failed to read any valid data from chunk files for epoch {epoch}. Final merged file cannot be created.")
            # Optionally clean up failed chunks? For now, leave them.
            _chunk_files = []
            _current_epoch = None
            return

        # Concatenate all dataframes
        print("Concatenating dataframes...")
        final_df = pd.concat(all_data_frames, ignore_index=True)

        # Verify and reorder columns just in case
        final_df = final_df.reindex(columns=_headers)

        # Determine final filename
        final_filename = _get_final_filename(epoch, _total_samples_logged_this_epoch)
        print(f"Saving final merged file to: {final_filename}")

        try:
            # Save the final merged dataframe
            final_df.to_excel(final_filename, index=False, engine='openpyxl')

            # Apply formatting to the final merged file
            _format_excel_file(final_filename)

            print(f"Successfully merged chunks into final file: {final_filename}")

            # 3. Clean up chunk files (optional)
            if _delete_chunks_after_merge:
                print("Cleaning up intermediate chunk files...")
                deleted_count = 0
                for chunk_file in _chunk_files:
                    try:
                        os.remove(chunk_file)
                        # print(f"Removed chunk file: {chunk_file}")
                        deleted_count += 1
                    except Exception as e:
                        print(f"Error removing chunk file {chunk_file}: {e}")
                print(f"Removed {deleted_count} chunk files.")
            else:
                print("Intermediate chunk files were kept.")

        except Exception as e:
            print(f"Error saving or formatting final merged file {final_filename}: {e}")

        finally:
            # 4. Reset state after finalization
            _chunk_files = []
            _buffer = []
            _samples_in_buffer = 0
            _total_samples_logged_this_epoch = 0
            _current_epoch = None  # Mark epoch as finalized
            print(f"Finalization process complete for epoch {epoch}.")


# Example Usage (for testing purposes, not part of the library integration)
if __name__ == '__main__':
    print("Running data_recorder example...")
    # Simulate usage across two epochs
    num_samples_epoch_1 = 25000
    num_samples_epoch_2 = 5000

    # --- Epoch 1 ---
    print("\n--- Starting Epoch 1 ---")
    set_epoch(1)
    for i in range(1, num_samples_epoch_1 + 1):
        mock_data_info = {'epoch': 1, 'seq_name': f'seq_{i}', 'template_ids': [f't{i}a', f't{i}b'],
                          'search_id': f's{i}'}
        mock_stats = {'Loss/total': random.random(), 'IoU': random.random() * 0.8}
        samples_stats_save(i, mock_data_info, mock_stats)
        if i % 5000 == 0:
            print(f"Logged sample {i}/{num_samples_epoch_1} for epoch 1")
    finalize_epoch(1)

    # --- Epoch 2 ---
    print("\n--- Starting Epoch 2 ---")
    set_epoch(2)
    for i in range(1, num_samples_epoch_2 + 1):
        mock_data_info = {'epoch': 2, 'seq_name': f'seq_{i + 100}', 'template_ids': [f't{i + 100}a', f't{i + 100}b'],
                          'search_id': f's{i + 100}'}
        mock_stats = {'Loss/total': random.random() * 0.5, 'IoU': random.random() * 0.9}
        samples_stats_save(i, mock_data_info, mock_stats)
        if i % 1000 == 0:
            print(f"Logged sample {i}/{num_samples_epoch_2} for epoch 2")
    finalize_epoch(2)

    print("\nData recorder example finished.")

import glob  # Import glob for file pattern matching


def reset_log():
    """Resets the data recorder state and deletes previous log files."""
    global _buffer, _chunk_files, _samples_in_buffer, _total_samples_logged_this_epoch, _current_epoch
    with _file_lock:
        print("Resetting data recorder state and deleting old log files...")
        # Reset global state variables
        _buffer = []
        _chunk_files = []
        _samples_in_buffer = 0
        _total_samples_logged_this_epoch = 0
        _current_epoch = None

        # Delete old log files matching the patterns
        deleted_count = 0
        # Use glob to find matching files in the current directory (.)
        chunk_pattern = "samples_log_epoch_*_sample_*.xlsx"
        final_pattern = "samples_log_epoch_*_all_sample_*.xlsx"

        files_to_delete = glob.glob(chunk_pattern) + glob.glob(final_pattern)

        if not files_to_delete:
            print("No old log files found to delete.")
        else:
            print(f"Found {len(files_to_delete)} old log files to delete:")
            for f in files_to_delete:
                try:
                    os.remove(f)
                    print(f"  - Deleted: {f}")
                    deleted_count += 1
                except OSError as e:
                    print(f"  - Error deleting file {f}: {e}")
            print(f"Finished deleting old log files. Total deleted: {deleted_count}")

        print("Data recorder reset complete.")


def save_gradients(model, sample_index, epoch, output_dir='gradients'):
    """
    Save model gradients to an HDF5 file.

    Args:
        model: The PyTorch model
        sample_index: Index of the current sample
        epoch: Current epoch number
        output_dir: Directory to save gradient files
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Create HDF5 file for this epoch if it doesn't exist
    h5_file = os.path.join(output_dir, f'gradients_epoch_{epoch}.h5')

    with h5py.File(h5_file, 'a') as f:  # 'a' mode allows appending to existing file
        # Create a group for this sample
        sample_grp = f.create_group(f'sample_{sample_index}')

        # Save gradients for each parameter
        for name, param in model.named_parameters():
            if param.grad is not None:
                # Convert gradient to numpy array and save
                grad_data = param.grad.detach().cpu().numpy()

                # Replace any problematic characters in parameter name that might cause issues with HDF5
                safe_name = name.replace('.', '_')

                # Save gradient data
                sample_grp.create_dataset(safe_name, data=grad_data, compression='gzip')

                # Also save some metadata
                sample_grp.attrs[f'{safe_name}_shape'] = str(param.grad.shape)
                sample_grp.attrs[f'{safe_name}_dtype'] = str(param.grad.dtype)

        # Save timestamp
        sample_grp.attrs['timestamp'] = time.strftime('%Y-%m-%d %H:%M:%S')
        sample_grp.attrs['sample_index'] = sample_index
        sample_grp.attrs['epoch'] = epoch